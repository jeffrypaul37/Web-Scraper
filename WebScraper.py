#web scraping a job site
#project by Jeffry Paul,Mohideen Irfan and Kesavan
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
url='https://www.monsterindia.com/search/portals-jobs?searchId=d69252c2-3e46-4987-9494-a589413c5f63'
agent = {"User-Agent":'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36'}
page = requests.get(url, headers=agent) 
soup = BeautifulSoup(page.text, 'html.parser')
print(soup)
w=openpyxl.Workbook() 
sheet=w.active
#set dimensions and make alignments
sheet.column_dimensions['A'].width=80
sheet.column_dimensions['B'].width=35
sheet.column_dimensions['C'].width=29
sheet.column_dimensions['D'].width=35
sheet.column_dimensions['E'].width=30
results=soup.find(id='srp-jobList')
jobs=results.find_all('div',class_='job-tittle')
i=2
sheet.cell(row=1,column=1).font=Font(size=25,italic=True,bold=True)
sheet.cell(row=1,column=2).font=Font(size=25,italic=True,bold=True)
sheet.cell(row=1,column=3).font=Font(size=25,italic=True,bold=True)
sheet.cell(row=1,column=4).font=Font(size=25,italic=True,bold=True)
sheet.cell(row=1,column=5).font=Font(size=25,italic=True,bold=True)
sheet.cell(row=1,column=1).alignment=Alignment(horizontal='center',vertical='center')
sheet.cell(row=1,column=2).alignment=Alignment(horizontal='center',vertical='center')
sheet.cell(row=1,column=3).alignment=Alignment(horizontal='center',vertical='center')
sheet.cell(row=1,column=4).alignment=Alignment(horizontal='center',vertical='center')
sheet.cell(row=1,column=5).alignment=Alignment(horizontal='center',vertical='center')
sheet.cell(row=1,column=1).value='JOB TITLE'
sheet.cell(row=1,column=2).value='COMPANY'
sheet.cell(row=1,column=3).value='LOCATION'
sheet.cell(row=1,column=4).value='VIEW LOCATION'
sheet.cell(row=1,column=5).value='APPLY LINK'
#get job title,company name, location and apply link from target webpage and print
for job in jobs:
    title=job.find('h3',class_='medium')
    company=job.find('span',class_='company-name')
    location=job.find('small')
    link=title.find('a', href=True)  
    if None in(title,company,location,link):
        continue
    print(title.text.strip())
    print(company.text.strip())
    print(location.text.strip())
    print ("APPLY LINK:", link['href'])
    print()
    print()
#print values in spreadsheet
    i=i+1
    sheet.cell(row=i,column=1).value=title.text.strip()
    sheet.cell(row=i,column=2).value=company.text.strip()
    sheet.cell(row=i,column=3).value=location.text.strip()
    #load links for job locations and applications
    sheet.cell(row=i,column=4).value = '=HYPERLINK("{}", "{}")'.format("https://www.google.com/maps/d/edit?mid=1UPY1xwhrraaPoAicLpG13gotjSvsKKiU&ll=33.088772001409325%2C83.1646314013193&z=4", "click here")
    sheet.cell(row=i,column=5).value=  '=HYPERLINK("{}", "{}")'.format(link['href'],"APPLY")
w.save('tech6.xlsx')
    






